from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_login import LoginManager, login_required, current_user, login_user, logout_user
from flask_migrate import Migrate
from models import db, Usuario, Categoria, Gasto, Ingreso, Empresa, Historial
# FASE-CAJA-GENERAL-S3: el vocabulario de "con que plata se pago" y las
# cuentas contra las que se valida. Se importan por nombre y no por modulo
# para no cambiar el estilo del resto del archivo.
from models import (
    ORIGENES_FONDO,
    ORIGEN_CAPITAL,
    ORIGEN_FACTURACION,
    CuentaCobro,
)
# `registrar_cambio` YA NO se importa de aca: la version de eva_utils usaba
# nombres de columna que no existen y nunca fallo solo porque el `def` de
# app.py:723 la pisaba. Se borro alla; la unica que queda es la de mas abajo.
from eva_utils import dias_de_operacion, generar_analisis_completo
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from flask_mail import Mail, Message
from functools import wraps
from sqlalchemy.orm import joinedload
import operator
import secrets
import string
import os
import sys
import io

# FORCE PRINT 
import sys
sys.stderr.write("🔍 INICIANDO APP - CARGANDO VARIABLES\n")
sys.stderr.flush()

# CONFIGURAR ENCODING UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# CARGAR VARIABLES DE ENTORNO
from dotenv import load_dotenv
import sys
sys.stderr.write("🔍 Intentando cargar .env...\n")
sys.stderr.flush()
load_dotenv()
sys.stderr.write(f"✓ MAIL_USERNAME: {os.environ.get('MAIL_USERNAME')}\n")
sys.stderr.flush()

# CREAR LA APP
app = Flask(__name__)
print("\n" + "="*60)
print("🚀 INICIANDO PYME - SISTEMA DE GESTIÓN FINANCIERA")
print("="*60)
SECRET_KEY = os.environ.get('SECRET_KEY')
if not SECRET_KEY:
    raise RuntimeError(
        "SECRET_KEY no esta definida. Configurala como variable de entorno "
        "antes de arrancar la aplicacion."
    )
app.config['SECRET_KEY'] = SECRET_KEY

# CIFRADO DE CREDENCIALES DE CANALES (FASE3-S1)
# Mismo criterio que SECRET_KEY: sin clave la app no arranca. El fallback
# silencioso seria guardar tokens de Tiendanube en texto plano.
import cripto
cripto.verificar_clave_configurada()

# BASE DE DATOS
# En produccion (Render) se define DATABASE_URL apuntando a Postgres.
# Sin esa variable, se cae al SQLite local para desarrollo.
DATABASE_URL = os.environ.get('DATABASE_URL')
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL or 'sqlite:///pyme.db'


def _ofuscar_uri(uri):
    """Devuelve la URI de conexion con la password enmascarada, para loguear."""
    import re
    return re.sub(r'://([^:/@]+):([^@]+)@', r'://\1:****@', uri)


print(f"🗄️  BASE DE DATOS: {_ofuscar_uri(app.config['SQLALCHEMY_DATABASE_URI'])}")
print(f"🗄️  MOTOR: {'PostgreSQL' if DATABASE_URL else 'SQLite (fallback local)'}")

# CONFIGURACIÓN DE EMAIL
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_USERNAME')

print(f"📧 MAIL_USERNAME: {app.config['MAIL_USERNAME']}")
print(f"🔐 MAIL_PASSWORD: {'*' * 20 if app.config['MAIL_PASSWORD'] else 'NO CONFIGURADO'}")
print("="*60 + "\n")

# INICIALIZAR MAIL Y DB
mail = Mail(app)
db.init_app(app)
migrate = Migrate(app, db)  # El esquema lo maneja Alembic, no db.create_all()

# INTEGRACIONES CON CANALES DE VENTA (FASE3-S1)
from rutas_integraciones import integraciones_bp
app.register_blueprint(integraciones_bp)

# CARGA MANUAL DE VENTAS PRESENCIALES (FASE3-S4)
from rutas_ventas import ventas_bp
app.register_blueprint(ventas_bp)

# LISTADO DE STOCK (FASE-STOCK-S1)
from rutas_productos import productos_bp
app.register_blueprint(productos_bp)

# MARGEN POR PRODUCTO Y CANAL (FASE-REPORTES-S3-MARGEN)
from rutas_reportes import reportes_bp
app.register_blueprint(reportes_bp)

# QUIEN HIZO QUE (FASE-AUDITORIA-S2)
# Se instala despues de los blueprints y no antes: el listener cuelga de
# db.session, no de las rutas, asi que cubre cualquier escritura de las tablas
# de su lista blanca venga de donde venga -- incluidas las que se agreguen
# manana sin que nadie se acuerde de auditarlas.
import auditoria
auditoria.instalar()


# El historial guarda `fecha` en UTC naive (datetime.utcnow, desde el dia uno).
# No se toca como se guarda -- reescribir las filas viejas seria inventar una
# hora que nadie registro -- se corrige al mostrarla. Argentina es UTC-3 todo
# el ano: no hay horario de verano desde 2009, asi que un offset fijo alcanza y
# evita meter una dependencia de zonas horarias.
HORAS_ARGENTINA = timedelta(hours=-3)


@app.template_filter('hora_local')
def hora_local(momento):
    """UTC naive -> hora de Argentina, lista para imprimir."""
    if momento is None:
        return ''
    return (momento + HORAS_ARGENTINA).strftime('%d/%m/%Y %H:%M')

@app.route('/')
def index():
    return redirect(url_for('login'))

# FUNCIONES DE EMAIL
def generar_codigo():
    return ''.join(secrets.choice(string.digits) for _ in range(6))

import threading

def enviar_email(destinatario, asunto, cuerpo):
    """Envía email de forma asincrónica en un thread separado"""
    def _enviar():
        try:
            print(f"\n🔍 Intentando enviar email a: {destinatario}")
            print(f"📧 Servidor: {app.config['MAIL_SERVER']}")
            print(f"👤 Usuario: {app.config['MAIL_USERNAME']}")
            
            msg = Message(
                asunto,
                recipients=[destinatario],
                html=cuerpo,
                charset='utf-8'
            )
            mail.send(msg)
            print(f"✓ Email enviado exitosamente a {destinatario}\n")
        except Exception as e:
            print(f"✗ Error al enviar email: {e}\n")
    
    # Ejecutar en un thread separado para no bloquear
    thread = threading.Thread(target=_enviar)
    thread.daemon = True
    thread.start()
    return True

# LOGIN MANAGER
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

def require_role(roles):
    """Protege rutas por rol de usuario"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('Debes iniciar sesión', 'danger')
                return redirect(url_for('login'))
            
            if current_user.rol not in roles:
                flash('No tienes permiso para acceder a esta página', 'danger')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Hacemos que Usuario sea compatible con Flask-Login
Usuario.is_authenticated = True
Usuario.is_active = True
Usuario.is_anonymous = False

def get_id(self):
    return str(self.id)

Usuario.get_id = get_id


# ======================== AUTENTICACIÓN ========================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        usuario = Usuario.query.filter_by(email=email).first()
        
        if usuario and usuario.check_password(password) and usuario.activo:
            # Validar que el email esté verificado
            if not usuario.verificado:
                flash('Por favor verifica tu email antes de iniciar sesión.', 'warning')
                return redirect(url_for('login'))
            
            login_user(usuario)
            flash(f'¡Bienvenido {usuario.nombre}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Email o contraseña incorrectos', 'danger')
    
    return render_template('login.html')

@app.route('/cambiar-password', methods=['GET', 'POST'])
def cambiar_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        
        usuario = Usuario.query.filter_by(email=email).first()
        if not usuario:
            flash('Email no encontrado', 'danger')
            return redirect(url_for('cambiar_password'))
        
        # Generar código
        codigo = generar_codigo()
        usuario.codigo_reset = codigo
        db.session.commit()
        
        # Enviar email
        cuerpo_email = f"""
        <h2>Verifica tu email</h2>
        <p>Tu código de verificacion es:</p>
        <h1 style="color: #1e90ff; letter-spacing: 5px;">{codigo}</h1>
        <p>Este código expira en 10 minutos.</p>
        """
        
        print(f"🔥 ANTES DE ENVIAR EMAIL - email: {email}")
        if enviar_email(email, "Código de Restablecimiento - PYME", cuerpo_email):
            session['email_reset'] = email
            flash('Se envio un código a tu email.', 'success')
            return redirect(url_for('verificar_codigo_reset'))
        else:
            flash('Error al enviar email. Intenta de nuevo.', 'danger')
    
    return render_template('cambiar_password.html')

@app.route('/verificar-codigo-reset', methods=['GET', 'POST'])
def verificar_codigo_reset():
    if 'email_reset' not in session:
        return redirect(url_for('cambiar_password'))
    
    if request.method == 'POST':
        codigo = request.form.get('codigo', '').strip()
        usuario = Usuario.query.filter_by(email=session['email_reset']).first()
        
        if usuario and codigo == usuario.codigo_reset:
            session['user_reset_id'] = usuario.id
            session.pop('email_reset', None)
            flash('Código verificado. Ingresa tu nueva contraseña.', 'success')
            return redirect(url_for('nueva_password'))
        else:
            flash('Codigo incorrecto.', 'danger')
    
    return render_template('verificar_codigo_reset.html', email=session.get('email_reset'))

@app.route('/nueva-password', methods=['GET', 'POST'])
def nueva_password():
    if 'user_reset_id' not in session:
        return redirect(url_for('cambiar_password'))
    
    if request.method == 'POST':
        password_nueva = request.form.get('password_nueva', '').strip()
        password_confirma = request.form.get('password_confirma', '').strip()
        
        if not password_nueva or len(password_nueva) < 6:
            flash('La contraseña debe tener minimo 6 caracteres.', 'danger')
            return redirect(url_for('nueva_password'))
        
        if password_nueva != password_confirma:
            flash('Las contraseñas no coinciden.', 'danger')
            return redirect(url_for('nueva_password'))
        
        usuario = Usuario.query.get(session['user_reset_id'])
        usuario.set_password(password_nueva)
        usuario.codigo_reset = None
        db.session.commit()
        
        session.pop('user_reset_id', None)
        
        flash('Contraseña actualizada correctamente. Inicia sesión.', 'success')
        return redirect(url_for('login'))
    
    return render_template('nueva_password.html')

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        nombre = request.form['nombre']
        email = request.form['email']
        password = request.form['password']
        
        if Usuario.query.filter_by(email=email).first():
            flash('El email ya esta registrado', 'error')
            return redirect(url_for('registro'))
        
        # Generar código de verificación
        codigo = generar_codigo()
        
        # Crear empresa
        empresa = Empresa(nombre=f"Empresa de {nombre}")
        db.session.add(empresa)
        db.session.flush()
        
        # Crear usuario sin verificar
        usuario = Usuario(nombre=nombre, email=email, empresa_id=empresa.id, verificado=False, codigo_verificacion=codigo)
        usuario.set_password(password)
        db.session.add(usuario)
        db.session.commit()
        
        # Enviar email
        cuerpo_email = f"""
        <h2>Verifica tu email</h2>
        <p>Tu codigo de verificacion es:</p>
        <h1 style="color: #1e90ff; letter-spacing: 5px;">{codigo}</h1>
        <p>Este codigo expira en 10 minutos.</p>
        """

        print(f"🔥 ANTES DE ENVIAR EMAIL - email: {email}")
        if enviar_email(email, "Codigo de Verificacion - PYME", cuerpo_email):
            session['user_id'] = usuario.id
            session['email_usuario'] = email
            flash('Se envio un codigo a tu email. Verifica tu cuenta.', 'success')
            return redirect(url_for('verificar_email'))
        else:
            db.session.delete(usuario)
            db.session.delete(empresa)
            db.session.commit()
            flash('Error al enviar email. Intenta de nuevo.', 'danger')
            return redirect(url_for('registro'))
    
    return render_template('registro.html')

@app.route('/verificar-email', methods=['GET', 'POST'])
def verificar_email():
    if 'user_id' not in session:
        return redirect(url_for('registro'))
    
    usuario = Usuario.query.get(session['user_id'])
    if not usuario:
        return redirect(url_for('registro'))
    
    if request.method == 'POST':
        codigo_ingresado = request.form.get('codigo', '').strip()
        
        if codigo_ingresado == usuario.codigo_verificacion:
            usuario.verificado = True
            usuario.codigo_verificacion = None
            db.session.commit()
            
            # Limpiar sesión temporal
            session.pop('user_id', None)
            session.pop('email_usuario', None)
            
            # Logear automáticamente al usuario
            login_user(usuario)
            
            flash('Email verificado correctamente. ¡Bienvenido!', 'success')
            return redirect(url_for('crear_empresa'))
        else:
            flash('Codigo incorrecto. Intenta de nuevo.', 'danger')
    
    return render_template('verificar_email.html', email=usuario.email)

@app.route('/crear-empresa', methods=['GET', 'POST'])
@login_required
def crear_empresa():
    usuario = current_user
    
    if request.method == 'POST':
        usuario.empresa.nombre = request.form['nombre']
        usuario.empresa.ruc = request.form.get('ruc')
        usuario.empresa.telefono = request.form.get('telefono')
        usuario.empresa.email = request.form.get('email')
        usuario.empresa.direccion = request.form.get('direccion')
        usuario.empresa.descripcion = request.form.get('descripcion')
        # FASE-EVA-S2: aca habia un
        #     usuario.empresa.capital_invertido = float(request.form.get('capital_invertido', 0))
        # que leia un campo que `crear_empresa.html` NUNCA renderizo. El
        # `.get(..., 0)` tapaba el KeyError, asi que la linea no fallaba: le
        # escribia 0 al capital en cada alta de empresa, prometiendo un dato que
        # el alta jamas pidio. Se saca la lectura muerta en vez de inventarle un
        # input al onboarding: el capital se carga en /config/eva, que es la
        # pantalla que existe para eso.
        db.session.commit()
        
        flash('¡Bienvenido! Tu empresa ha sido creada', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('crear_empresa.html', usuario=usuario)

@app.route('/logout', methods=['POST'])
@login_required
def logout():
    logout_user()
    flash('Sesión cerrada', 'success')
    return redirect(url_for('login'))

# ================= HELPERS DE GASTO / INGRESO / CATEGORIA =================
# FASE-CAJA-GENERAL-S2. Los tres nacen de la misma correccion: gasto e ingreso
# pasaron a ser de la EMPRESA y no del usuario, asi que todo lo que los rodea
# -- el rango de fechas, el combo de categorias, la fecha del alta -- tuvo que
# dejar de mirar `current_user.id`.

def _filtrar_por_fecha(consulta, columna):
    """Aplica el rango ?fecha_inicio / ?fecha_fin a una consulta, EN SQL.

    Antes cada listado traia todas las filas de la tabla y filtraba la lista
    en Python. Con dos gastos daba igual; con el libro de caja de un anio, no.

    Una fecha que no parsea se ignora en vez de voltear el request con un 500:
    el filtro es una comodidad de la pantalla, no un dato que valga romper.
    """
    for clave, comparar in (('fecha_inicio', operator.ge),
                            ('fecha_fin', operator.le)):
        crudo = (request.args.get(clave) or '').strip()
        if not crudo:
            continue
        try:
            limite = datetime.strptime(crudo, '%Y-%m-%d').date()
        except ValueError:
            continue
        consulta = consulta.filter(comparar(columna, limite))
    return consulta


def _fecha_del_form(crudo, por_defecto=None):
    """El campo `fecha` del form como date. Invalido = ValueError.

    Vacio devuelve `por_defecto`, o hoy si no se paso ninguno. Las ediciones
    pasan la fecha que ya tenia la fila: un campo que no vino no puede
    significar "movela a hoy" en silencio.

    POR QUE EXISTE

    El alta hardcodeaba `datetime.now().date()` y el form ni ofrecia el campo:
    era imposible cargar el historico del Excel, que arranca el 28/07. Y la
    edicion hacia `request.form['fecha']` contra una plantilla que no tenia
    ese input -- o sea que editar un gasto SIEMPRE reventaba con KeyError, y
    el `except Exception` de la ruta lo mostraba como "Error al editar: fecha".
    Nadie lo noto porque las tablas estaban vacias.

    `.get()` en vez de `[]` es la mitad del arreglo; la otra mitad es el input
    en las cuatro plantillas.
    """
    crudo = (crudo or '').strip()
    if not crudo:
        return por_defecto or datetime.now().date()
    return datetime.strptime(crudo, '%Y-%m-%d').date()


def _categorias_de(empresa_id, tipo=None):
    """Las categorias de la EMPRESA, no las del usuario logueado.

    Hasta FASE-CATEGORIA-S1 esto era un JOIN a `usuario`: `categoria` colgaba
    de `usuario_id` y la unica forma de llegar a la empresa era pasando por la
    persona que la habia tipeado. Funcionaba, pero era un parche sobre un
    esquema equivocado -- y se caia con la fila cuyo usuario ya no existe.
    Ahora `categoria.empresa_id` dice el duenio directamente y esto es un
    filtro sobre una columna indexada.

    La funcion queda porque centraliza las otras dos decisiones (el filtro por
    tipo y el orden alfabetico) que los diez call sites comparten.
    """
    consulta = Categoria.query.filter_by(empresa_id=empresa_id)
    if tipo:
        consulta = consulta.filter(Categoria.tipo == tipo)
    return consulta.order_by(Categoria.nombre)


def _cuentas_de(empresa_id):
    """Las cuentas de cobro de la EMPRESA, para el selector de "de que cuenta".

    Hoy son dos: la de Roman y la de Nachi (ids 40 y 41). No se filtra por
    `activo` a proposito: si manana una cuenta se apaga, los gastos ya
    cargados contra ella tienen que poder seguir editandose sin que la opcion
    desaparezca del <select> y el guardado se lleve puesto el dato.
    """
    return (CuentaCobro.query
            .filter_by(empresa_id=empresa_id)
            .order_by(CuentaCobro.id)
            .all())


def _origen_del_form(form, empresa_id, requerido=False,
                     por_defecto=(None, None)):
    """FASE-CAJA-GENERAL-S3: lee "con que plata se pago" -> (origen, cuenta_id).

    Levanta ValueError con el texto que se le muestra a quien carga. Es la
    unica puerta por la que el par (origen_fondo, cuenta_pago_id) entra al
    modelo desde una pantalla, asi que las dos reglas del CHECK se aplican aca
    antes de que la base tenga que defenderse.

    `requerido` es True en el alta y False en la edicion, y la diferencia es
    la misma que ya distingue `_fecha_del_form`: un campo que NO vino no puede
    significar un cambio destructivo. En la edicion, sin origen elegido se
    devuelve `por_defecto` -- lo que la fila ya decia -- asi que:

      - un gasto viejo con origen_fondo NULL se sigue pudiendo editar sin que
        nadie tenga que inventar de que bolsillo salio, y
      - uno que si lo tiene no lo pierde porque el formulario haya llegado
        incompleto.

    En el alta no hay nada previo que conservar y por eso ahi si es
    obligatorio: un gasto nuevo que no dice de que plata salio deja el saldo
    real de los socios mintiendo por omision desde el primer dia.
    """
    origen = (form.get('origen_fondo') or '').strip() or None

    if origen is None:
        if requerido:
            raise ValueError('Decí con qué plata se pagó: facturación o capital')
        return por_defecto

    if origen not in ORIGENES_FONDO:
        raise ValueError('El origen del gasto no es válido')

    if origen == ORIGEN_CAPITAL:
        # La cuenta que venga se DESCARTA en silencio, no se rechaza. El
        # <select> de cuenta sigue en el DOM cuando se elige Capital -- se
        # oculta, no se borra -- asi que el navegador manda igual lo que
        # estuviera seleccionado antes de cambiar de opcion. Rechazarlo seria
        # devolverle un error a quien carga por algo que hizo el formulario
        # solo; el dato que importa ("salio de capital") ya quedo dicho, y el
        # CHECK de la base sigue cubriendo cualquier otro camino de escritura.
        return ORIGEN_CAPITAL, None

    crudo = (form.get('cuenta_pago_id') or '').strip()
    if not crudo:
        raise ValueError('Si el gasto salió de la facturación, decí de qué '
                         'cuenta salió')

    try:
        cuenta_id = int(crudo)
    except ValueError:
        cuenta = None
    else:
        # Por EMPRESA, igual que la categoria: nadie paga desde la cuenta de
        # otra empresa aunque escriba el id a mano.
        cuenta = CuentaCobro.query.filter_by(id=cuenta_id,
                                             empresa_id=empresa_id).first()
    if cuenta is None:
        raise ValueError('La cuenta de pago no es válida')

    return ORIGEN_FACTURACION, cuenta.id


# ======================== DASHBOARD ========================

@app.route('/dashboard')
@login_required
def dashboard():
    # FILTRAR POR EMPRESA (SIN FILTRO DE FECHA)
    empresa_id = current_user.empresa_id
    gastos = Gasto.query.filter_by(empresa_id=empresa_id).all()
    ingresos = Ingreso.query.filter_by(empresa_id=empresa_id).all()
    
    config = current_user.empresa
    
    # eva_utils opera con floats (tasa_impuestos/capital_invertido son Float),
    # asi que convertimos los totales Decimal en el borde para no mezclar tipos.
    total_gastos = float(sum(g.monto for g in gastos))
    total_ingresos = float(sum(i.monto for i in ingresos))

    # FASE-EVA-S3: sobre cuantos dias se estan sumando esos dos totales.
    #
    # Es el MIN(fecha) de las dos tablas, y sale de las listas que ya estan en
    # memoria dos lineas mas arriba: hacerlo con un SELECT MIN aparte serian dos
    # consultas nuevas para releer filas que ya trajimos. `dias_de_operacion`
    # devuelve None si no hay ninguna, que es el caso que el estado neutral de
    # S2 ya cubre.
    fechas = [g.fecha for g in gastos] + [i.fecha for i in ingresos]
    dias = dias_de_operacion(min(fechas) if fechas else None)

    analisis = generar_analisis_completo(total_ingresos, total_gastos, config,
                                         dias)
    gastos_cat = gastos_por_categoria(gastos)
    ingresos_cat = ingresos_por_categoria(ingresos)
    
    # Últimos movimientos
    gastos_recientes = Gasto.query.filter_by(empresa_id=empresa_id).order_by(Gasto.fecha.desc()).limit(10).all()
    ingresos_recientes = Ingreso.query.filter_by(empresa_id=empresa_id).order_by(Ingreso.fecha.desc()).limit(10).all()
    
    return render_template('dashboard.html',
                         current_user=current_user,
                         usuario=current_user,
                         analisis=analisis,
                         gastos_cat=gastos_cat,
                         ingresos_cat=ingresos_cat,
                         gastos_recientes=gastos_recientes,
                         ingresos_recientes=ingresos_recientes)

# ======================== GASTOS ========================

@app.route('/gasto/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_gasto():
    if request.method == 'POST':
        try:
            descripcion = request.form.get('descripcion', '').strip()
            monto_str = request.form.get('monto', '').strip()

            # Validaciones
            if not descripcion:
                flash('La descripción es obligatoria', 'danger')
                return redirect(url_for('nuevo_gasto'))

            if not monto_str:
                flash('El monto es obligatorio', 'danger')
                return redirect(url_for('nuevo_gasto'))

            try:
                monto = Decimal(monto_str)
            except InvalidOperation:
                flash('El monto debe ser un número válido', 'danger')
                return redirect(url_for('nuevo_gasto'))

            if monto <= 0:
                flash('El monto debe ser mayor a 0', 'danger')
                return redirect(url_for('nuevo_gasto'))

            # La fecha la elige quien carga. Vacia = hoy; invalida se avisa en
            # vez de guardar el gasto en un dia que nadie pidio.
            try:
                fecha = _fecha_del_form(request.form.get('fecha'))
            except ValueError:
                flash('La fecha no es válida (formato AAAA-MM-DD)', 'danger')
                return redirect(url_for('nuevo_gasto'))

            if not request.form.get('categoria_id'):
                flash('Debes seleccionar una categoría', 'danger')
                return redirect(url_for('nuevo_gasto'))

            # La categoría tiene que ser de la EMPRESA (ver _categorias_de)
            categoria = _categorias_de(current_user.empresa_id).filter(
                Categoria.id == int(request.form['categoria_id'])).first()
            if not categoria:
                flash('Categoría inválida', 'danger')
                return redirect(url_for('nuevo_gasto'))

            # FASE-CAJA-GENERAL-S3: de que plata salio. Obligatorio en el
            # alta -- un gasto nuevo que no dice de que bolsillo salio deja
            # el saldo real de los socios mintiendo por omision desde el
            # primer dia, y completarlo despues es adivinar.
            try:
                origen_fondo, cuenta_pago_id = _origen_del_form(
                    request.form, current_user.empresa_id, requerido=True)
            except ValueError as e:
                flash(str(e), 'danger')
                return redirect(url_for('nuevo_gasto'))

            gasto = Gasto(
                fecha=fecha,
                descripcion=descripcion,
                monto=monto,
                categoria_id=categoria.id,
                empresa_id=current_user.empresa_id,
                usuario_id=current_user.id,
                origen_fondo=origen_fondo,
                cuenta_pago_id=cuenta_pago_id
            )
            db.session.add(gasto)
            db.session.commit()

            registrar_cambio(current_user.id, 'crear', 'gasto', gasto.id,
                           f'Gasto de ${monto} - {descripcion}')

            flash('✅ Gasto registrado exitosamente', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash(f'❌ Error al agregar gasto: {str(e)}', 'danger')

    # CATEGORÍAS DE LA EMPRESA
    categorias = _categorias_de(current_user.empresa_id, tipo='gasto').all()
    return render_template('agregar_gasto.html', categorias=categorias,
                           cuentas=_cuentas_de(current_user.empresa_id),
                           origenes=ORIGENES_FONDO,
                           origen_facturacion=ORIGEN_FACTURACION,
                           hoy=datetime.now().date().isoformat())

@app.route('/gasto/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_gasto(id):
    gasto = Gasto.query.get_or_404(id)

    # Verificar que pertenezca a la EMPRESA
    if gasto.empresa_id != current_user.empresa_id:
        flash('No tienes permiso', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        try:
            descripcion = request.form.get('descripcion', '').strip()
            monto = Decimal(request.form.get('monto') or 0)

            if not descripcion or monto <= 0:
                flash('Complete todos los campos correctamente', 'danger')
                return redirect(url_for('editar_gasto', id=id))

            try:
                fecha = _fecha_del_form(request.form.get('fecha'),
                                        por_defecto=gasto.fecha)
            except ValueError:
                flash('La fecha no es válida (formato AAAA-MM-DD)', 'danger')
                return redirect(url_for('editar_gasto', id=id))

            categoria = _categorias_de(current_user.empresa_id).filter(
                Categoria.id == int(request.form.get('categoria_id') or 0)).first()
            if not categoria:
                flash('Categoría inválida', 'danger')
                return redirect(url_for('editar_gasto', id=id))

            # FASE-CAJA-GENERAL-S3: si el form no trae origen, se deja el
            # que la fila ya tenia (que puede ser NULL). Es la misma regla
            # que la fecha dos lineas mas arriba: un campo que no vino no
            # significa "borralo".
            try:
                origen_fondo, cuenta_pago_id = _origen_del_form(
                    request.form, current_user.empresa_id,
                    por_defecto=(gasto.origen_fondo, gasto.cuenta_pago_id))
            except ValueError as e:
                flash(str(e), 'danger')
                return redirect(url_for('editar_gasto', id=id))

            gasto.descripcion = descripcion
            gasto.monto = monto
            gasto.categoria_id = categoria.id
            gasto.fecha = fecha
            gasto.origen_fondo = origen_fondo
            gasto.cuenta_pago_id = cuenta_pago_id

            db.session.commit()

            registrar_cambio(current_user.id, 'editar', 'gasto', gasto.id,
                           f'Gasto actualizado - {descripcion}')

            flash('Gasto actualizado', 'success')
            return redirect(url_for('listar_gastos'))
        except Exception as e:
            flash(f'Error al editar: {str(e)}', 'danger')

    # CATEGORÍAS DE LA EMPRESA
    categorias = _categorias_de(current_user.empresa_id, tipo='gasto').all()
    return render_template('editar_gasto.html', gasto=gasto,
                           categorias=categorias,
                           cuentas=_cuentas_de(current_user.empresa_id),
                           origenes=ORIGENES_FONDO,
                           origen_facturacion=ORIGEN_FACTURACION)

@app.route('/gasto/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_gasto(id):
    gasto = Gasto.query.get_or_404(id)

    # Verificar que pertenezca a la EMPRESA
    if gasto.empresa_id != current_user.empresa_id:
        flash('No tienes permiso', 'danger')
        return redirect(url_for('dashboard'))

    descripcion = gasto.descripcion
    monto = gasto.monto

    db.session.delete(gasto)
    db.session.commit()

    registrar_cambio(current_user.id, 'eliminar', 'gasto', id,
                   f'Gasto eliminado - ${monto} - {descripcion}')

    flash('Gasto eliminado', 'success')
    return redirect(url_for('listar_gastos'))

@app.route('/gasto/listar')
@login_required
def listar_gastos():
    # POR EMPRESA, y el rango de fechas EN SQL (ver _filtrar_por_fecha)
    gastos = (_filtrar_por_fecha(
                  Gasto.query.filter_by(empresa_id=current_user.empresa_id),
                  Gasto.fecha)
              .order_by(Gasto.fecha.desc())
              .all())

    return render_template('listar_gastos.html', gastos=gastos)

# ======================== INGRESOS ========================

@app.route('/ingreso/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_ingreso():
    if request.method == 'POST':
        try:
            descripcion = request.form.get('descripcion', '').strip()
            monto_str = request.form.get('monto', '').strip()

            # Las validaciones son las mismas que las de nuevo_gasto, campo por
            # campo. Antes esta ruta hacia `int(request.form['categoria_id'])`
            # sin mirar nada: sin categoría elegida reventaba con KeyError y el
            # `except Exception` de abajo lo mostraba como "Error al agregar
            # ingreso: 'categoria_id'", que no le dice a nadie qué falta.
            if not descripcion:
                flash('La descripción es obligatoria', 'danger')
                return redirect(url_for('nuevo_ingreso'))

            if not monto_str:
                flash('El monto es obligatorio', 'danger')
                return redirect(url_for('nuevo_ingreso'))

            try:
                monto = Decimal(monto_str)
            except InvalidOperation:
                flash('El monto debe ser un número válido', 'danger')
                return redirect(url_for('nuevo_ingreso'))

            if monto <= 0:
                flash('El monto debe ser mayor a 0', 'danger')
                return redirect(url_for('nuevo_ingreso'))

            try:
                fecha = _fecha_del_form(request.form.get('fecha'))
            except ValueError:
                flash('La fecha no es válida (formato AAAA-MM-DD)', 'danger')
                return redirect(url_for('nuevo_ingreso'))

            if not request.form.get('categoria_id'):
                flash('Debes seleccionar una categoría', 'danger')
                return redirect(url_for('nuevo_ingreso'))

            # La categoría tiene que ser de la EMPRESA (ver _categorias_de)
            categoria = _categorias_de(current_user.empresa_id).filter(
                Categoria.id == int(request.form['categoria_id'])).first()
            if not categoria:
                flash('Categoría inválida', 'danger')
                return redirect(url_for('nuevo_ingreso'))

            ingreso = Ingreso(
                fecha=fecha,
                descripcion=descripcion,
                monto=monto,
                categoria_id=categoria.id,
                empresa_id=current_user.empresa_id,
                usuario_id=current_user.id
            )
            db.session.add(ingreso)
            db.session.commit()

            registrar_cambio(current_user.id, 'crear', 'ingreso', ingreso.id,
                           f'Ingreso de ${monto} - {descripcion}')

            flash('Ingreso agregado exitosamente', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash(f'Error al agregar ingreso: {str(e)}', 'danger')

    # CATEGORÍAS DE LA EMPRESA
    categorias = _categorias_de(current_user.empresa_id, tipo='ingreso').all()
    return render_template('agregar_ingreso.html', categorias=categorias,
                           hoy=datetime.now().date().isoformat())

@app.route('/ingreso/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_ingreso(id):
    ingreso = Ingreso.query.get_or_404(id)

    # Verificar que pertenezca a la EMPRESA
    if ingreso.empresa_id != current_user.empresa_id:
        flash('No tienes permiso', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        try:
            descripcion = request.form.get('descripcion', '').strip()
            monto = Decimal(request.form.get('monto') or 0)

            if not descripcion or monto <= 0:
                flash('Complete todos los campos correctamente', 'danger')
                return redirect(url_for('editar_ingreso', id=id))

            try:
                fecha = _fecha_del_form(request.form.get('fecha'),
                                        por_defecto=ingreso.fecha)
            except ValueError:
                flash('La fecha no es válida (formato AAAA-MM-DD)', 'danger')
                return redirect(url_for('editar_ingreso', id=id))

            categoria = _categorias_de(current_user.empresa_id).filter(
                Categoria.id == int(request.form.get('categoria_id') or 0)).first()
            if not categoria:
                flash('Categoría inválida', 'danger')
                return redirect(url_for('editar_ingreso', id=id))

            ingreso.descripcion = descripcion
            ingreso.monto = monto
            ingreso.categoria_id = categoria.id
            ingreso.fecha = fecha

            db.session.commit()

            registrar_cambio(current_user.id, 'editar', 'ingreso', ingreso.id,
                           f'Ingreso actualizado - {descripcion}')

            flash('Ingreso actualizado', 'success')
            return redirect(url_for('listar_ingresos'))
        except Exception as e:
            flash(f'Error al editar: {str(e)}', 'danger')

    # CATEGORÍAS DE LA EMPRESA
    categorias = _categorias_de(current_user.empresa_id, tipo='ingreso').all()
    return render_template('editar_ingreso.html', ingreso=ingreso, categorias=categorias)

@app.route('/ingreso/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_ingreso(id):
    ingreso = Ingreso.query.get_or_404(id)

    # Verificar que pertenezca a la EMPRESA
    if ingreso.empresa_id != current_user.empresa_id:
        flash('No tienes permiso', 'danger')
        return redirect(url_for('dashboard'))

    descripcion = ingreso.descripcion
    monto = ingreso.monto

    db.session.delete(ingreso)
    db.session.commit()

    registrar_cambio(current_user.id, 'eliminar', 'ingreso', id,
                   f'Ingreso eliminado - ${monto} - {descripcion}')

    flash('Ingreso eliminado', 'success')
    return redirect(url_for('listar_ingresos'))

@app.route('/ingreso/listar')
@login_required
def listar_ingresos():
    # POR EMPRESA, y el rango de fechas EN SQL (ver _filtrar_por_fecha)
    ingresos = (_filtrar_por_fecha(
                    Ingreso.query.filter_by(empresa_id=current_user.empresa_id),
                    Ingreso.fecha)
                .order_by(Ingreso.fecha.desc())
                .all())

    return render_template('listar_ingresos.html', ingresos=ingresos)

# ======================== CAJA GENERAL ========================

# El cero con el que arranca el saldo, y el tipo de todos los montos del libro.
# Decimal, nunca float: son pesos.
CERO = Decimal('0.00')


@app.route('/caja-general')
@login_required
def caja_general():
    """El libro unico de la empresa: gasto e ingreso en una sola lista, con
    saldo corriente. Es la hoja CAJA GENERAL del Excel de Roman.

    LA CUENTA

    Es la columna E del Excel, `=SUM(E_anterior + entrada - salida)`: el saldo
    de cada fila es el de la anterior mas lo que entro menos lo que salio. Se
    acumula una sola vez recorriendo la lista, no se recalcula la suma entera
    en cada fila -- que da el mismo numero pero es cuadratico, y este libro
    esta hecho para crecer todos los dias.

    El saldo arranca en 0. No hay saldo inicial en el modelo, y el Excel
    tampoco lo tiene: su primera fila ya baja a -969.285 porque la mercaderia
    inicial se cargo como salida y la plata que Roman puso no se cargo como
    nada. Con el aporte de capital cargado como ingreso (categoria "Aporte de
    capital (socios)"), el arranque en 0 pasa a ser el correcto y no una
    coincidencia.

    LO QUE NO HACE

    No mira un solo pedido. Las ventas entran a mano, como ingreso, igual que
    en el Excel -- sumar `pedido.total` ademas de eso contaria la misma plata
    dos veces, y el Excel usa el neto por venta mientras que `pedido.total` es
    el bruto. Que esas dos cosas se junten es una decision para otra slice.
    """
    empresa_id = current_user.empresa_id

    ingresos = (Ingreso.query
                .options(joinedload(Ingreso.categoria))
                .filter_by(empresa_id=empresa_id)
                .all())
    gastos = (Gasto.query
              # FASE-CAJA-GENERAL-S3: `cuenta_pago` entra al joinedload porque
              # `origen_legible` la lee para escribir "Facturación — Roman".
              # Sin esto la pantalla dispara una consulta por cada gasto.
              .options(joinedload(Gasto.categoria),
                       joinedload(Gasto.cuenta_pago))
              .filter_by(empresa_id=empresa_id)
              .all())

    movimientos = [
        {'tipo': 'ingreso', 'id': i.id, 'fecha': i.fecha,
         'descripcion': i.descripcion,
         'categoria': i.categoria.nombre if i.categoria else None,
         # El ingreso no tiene origen: es plata que ENTRA, y de donde entro ya
         # lo dice su categoria. La clave existe igual para que la plantilla
         # no tenga que preguntar por el tipo antes de leerla.
         'origen': None, 'origen_puesto': True,
         'entrada': i.monto, 'salida': CERO}
        for i in ingresos
    ] + [
        {'tipo': 'gasto', 'id': g.id, 'fecha': g.fecha,
         'descripcion': g.descripcion,
         'categoria': g.categoria.nombre if g.categoria else None,
         'origen': g.origen_legible,
         # Aparte de la etiqueta, si la fila lo dice o no: la plantilla pinta
         # distinto el "sin dato" para que el hueco se vea, y comparar contra
         # el texto de la etiqueta seria atarla a como esta escrito.
         'origen_puesto': g.origen_fondo is not None,
         'entrada': CERO, 'salida': g.monto}
        for g in gastos
    ]

    # Fecha, despues id. `tipo` cierra el desempate porque gasto e ingreso
    # tienen secuencias de id separadas: sin el, dos filas del mismo dia con
    # el mismo id quedarian en un orden que depende del planificador y el
    # saldo se leeria distinto en cada recarga.
    movimientos.sort(key=lambda m: (m['fecha'], m['id'], m['tipo']))

    saldo = CERO
    for movimiento in movimientos:
        saldo += movimiento['entrada'] - movimiento['salida']
        movimiento['saldo'] = saldo

    return render_template(
        'caja_general.html',
        movimientos=movimientos,
        total_entradas=sum((m['entrada'] for m in movimientos), CERO),
        total_salidas=sum((m['salida'] for m in movimientos), CERO),
        saldo_final=saldo)


# ======================== CONFIGURACIÓN ========================

@app.route('/config/eva', methods=['GET', 'POST'])
@login_required
def config_eva():
    """Configuracion de la empresa: los datos y los tres parametros del EVA.

    FASE-EVA-S2: la plantilla POSTEA `nombre`, `ruc` y `email` desde la seccion
    "Datos de la Empresa" -- con los valores actuales precargados y `nombre`
    marcado `required` -- y esta ruta escribia SOLO los tres campos del EVA. O
    sea que corregir el nombre de la empresa mostraba "Configuracion
    actualizada" y no guardaba nada.

    De las dos salidas posibles se eligio conectar el guardado, no sacar los
    campos del form, porque sacarlos dejaba a Korvo sin ninguna forma de editar
    el nombre: el otro lugar donde se escriben esos tres campos es
    `crear_empresa()`, y a /crear-empresa se llega una sola vez, en el redirect
    de la verificacion de email (app.py:380). No hay link a esa pantalla desde
    ningun menu. Sacar los inputs habria arreglado la mentira perdiendo la
    unica via de correccion; conectarlos son tres lineas y no pierde nada.
    """
    empresa = current_user.empresa

    if request.method == 'POST':
        try:
            nombre = request.form.get('nombre', '').strip()
            if not nombre:
                flash('El nombre de la empresa es obligatorio', 'danger')
                return redirect(url_for('config_eva'))

            empresa.nombre = nombre
            # `or None` para que borrar el campo deje NULL y no un string vacio:
            # la plantilla ya hace `empresa.ruc or ''`, y un '' guardado se leia
            # igual que un NULL pero ensuciaba la columna.
            empresa.ruc = request.form.get('ruc', '').strip() or None
            empresa.email = request.form.get('email', '').strip() or None

            empresa.tasa_costo_capital = float(request.form['tasa_costo_capital'])
            empresa.capital_invertido = float(request.form['capital_invertido'])
            empresa.tasa_impuestos = float(request.form['tasa_impuestos']) / 100

            db.session.commit()

            registrar_cambio(current_user.id, 'editar', 'configuracion', empresa.id,
                           'Configuración de empresa actualizada')

            flash('Configuración actualizada', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')

    return render_template('configurar_eva.html', empresa=empresa)

# ======================== HISTORIAL ========================

@app.route('/historial')
@login_required
def ver_historial():
    # FILTRAR POR EMPRESA, no por usuario (FASE-AUDITORIA-S2).
    #
    # Filtraba por `usuario_id=current_user.id`, o sea que cada uno veia solo
    # lo suyo. Con un solo usuario daba igual; el dia que Nachi tenga login,
    # Roman no veria una sola de sus acciones -- justo lo contrario de lo que
    # la pantalla existe para mostrar.
    #
    # Por empresa tambien entran las filas de las cuentas ya eliminadas
    # (usuario_id en NULL): son las que mas importa no perder.
    historial = (Historial.query
                 .filter_by(empresa_id=current_user.empresa_id)
                 .order_by(Historial.fecha.desc())
                 .all())
    return render_template('historial.html', historial=historial, usuario=current_user)

@app.route('/historial/limpiar', methods=['POST'])
@login_required
def limpiar_historial():
    # YA NO BORRA NADA (FASE-AUDITORIA-S2).
    #
    # Antes hacia `.delete()` sobre el historial propio: un registro que el
    # auditado puede borrar de un boton no audita nada. Ahora la ruta deja
    # constancia de que alguien quiso limpiarlo y sigue de largo.
    #
    # La fila se escribe con `registrar_cambio`, la misma de siempre. El hook
    # de auditoria.py no la cubre: Historial no esta -- ni puede estar -- en su
    # lista blanca, porque auditarse a si misma seria recursion infinita.
    registrar_cambio(current_user.id, 'eliminar', 'historial', None,
                     'Se pidio limpiar el historial (no se borro ninguna fila)')

    flash('El historial ya no se borra: queda constancia de quién lo pidió y cuándo.',
          'warning')
    return redirect(url_for('ver_historial'))

def registrar_cambio(usuario_id, accion, tipo, id_registro, descripcion):
    """Escritura manual de historial. La usan las cuatro entidades viejas
    (gasto, ingreso, categoria, configuracion); el resto lo cubre el hook de
    auditoria.py sin que nadie tenga que llamar nada.

    `empresa_id` se deriva del usuario: es NOT NULL y ninguno de los once call
    sites lo pasa, y no se tocan.
    """
    usuario = Usuario.query.get(usuario_id)
    cambio = Historial(
        usuario_id=usuario_id,
        empresa_id=usuario.empresa_id if usuario else None,
        accion=accion,
        tipo=tipo,
        id_registro=id_registro,
        descripcion=descripcion
    )
    db.session.add(cambio)
    db.session.commit()

# ======================== CATEGORÍAS PERSONALIZABLES ========================

@app.route('/nueva-categoria', methods=['GET', 'POST'])
@login_required
def nueva_categoria():
    if request.method == 'POST':
        try:
            nombre = request.form.get('nombre', '').strip()
            tipo = request.form.get('tipo', '').strip()
            
            if not nombre:
                flash('El nombre es obligatorio', 'danger')
                return redirect(url_for('nueva_categoria'))
            
            if not tipo or tipo not in ['gasto', 'ingreso']:
                flash('Debes seleccionar un tipo valido', 'danger')
                return redirect(url_for('nueva_categoria'))
            
            if _categorias_de(current_user.empresa_id).filter(
                    Categoria.nombre == nombre).first():
                flash('Esta categoria ya existe', 'danger')
                return redirect(url_for('nueva_categoria'))
            
            categoria = Categoria(nombre=nombre, tipo=tipo,
                                  empresa_id=current_user.empresa_id,
                                  usuario_id=current_user.id)
            db.session.add(categoria)
            db.session.commit()
            
            registrar_cambio(current_user.id, 'crear', 'categoria', categoria.id, 
                           f'Categoria {nombre} creada')
            
            flash('Categoria creada exitosamente', 'success')
            return redirect(url_for('nueva_categoria'))
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
            db.session.rollback()
    
    return render_template('nueva_categoria.html')

@app.route('/categorias')
@login_required
def listar_categorias():
    categorias = _categorias_de(current_user.empresa_id).all()
    
    # Contar transacciones por categoría
    categorias_datos = []
    for cat in categorias:
        gastos = Gasto.query.filter_by(categoria_id=cat.id).count()
        ingresos = Ingreso.query.filter_by(categoria_id=cat.id).count()
        total_transacciones = gastos + ingresos
        
        categorias_datos.append({
            'categoria': cat,
            'total_transacciones': total_transacciones,
            'gastos': gastos,
            'ingresos': ingresos
        })
    
    return render_template('listar_categorias.html', categorias_datos=categorias_datos)

@app.route('/categoria/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_categoria(id):
    categoria = Categoria.query.get_or_404(id)

    # Verificar que pertenezca a la EMPRESA
    if categoria.empresa_id != current_user.empresa_id:
        flash('No tienes permiso', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            nombre = request.form.get('nombre', '').strip()
            
            if not nombre:
                flash('El nombre es obligatorio', 'danger')
                return redirect(url_for('editar_categoria', id=id))
            
            categoria.nombre = nombre
            db.session.commit()
            
            registrar_cambio(current_user.id, 'editar', 'categoria', categoria.id, 
                           f'Categoría actualizada - {nombre}')
            
            flash('Categoría actualizada', 'success')
            return redirect(url_for('listar_categorias'))
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('editar_categoria.html', categoria=categoria)

@app.route('/categoria/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_categoria(id):
    categoria = Categoria.query.get_or_404(id)

    # Verificar que pertenezca a la EMPRESA
    if categoria.empresa_id != current_user.empresa_id:
        flash('No tienes permiso', 'danger')
        return redirect(url_for('dashboard'))
    
    # Verificar si está en uso
    if Gasto.query.filter_by(categoria_id=id).first() or Ingreso.query.filter_by(categoria_id=id).first():
        flash('No puedes eliminar una categoría que está en uso', 'danger')
        return redirect(url_for('listar_categorias'))
    
    nombre = categoria.nombre
    db.session.delete(categoria)
    db.session.commit()
    
    registrar_cambio(current_user.id, 'eliminar', 'categoria', id, f'Categoría {nombre} eliminada')
    
    flash('Categoría eliminada', 'success')
    return redirect(url_for('listar_categorias'))

# ======================== EXPORTAR A EXCEL ========================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file

@app.route('/exportar/gastos')
@login_required
def exportar_gastos():

    gastos = _filtrar_por_fecha(
        Gasto.query.filter_by(empresa_id=current_user.empresa_id),
        Gasto.fecha).order_by(Gasto.fecha).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"
    
    # Estilos
    header_fill = PatternFill(start_color="1e90ff", end_color="1e90ff", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    headers = ['Fecha', 'Descripción', 'Categoría', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, gasto in enumerate(gastos, 2):
        ws.cell(row=row, column=1).value = gasto.fecha
        ws.cell(row=row, column=2).value = gasto.descripcion
        ws.cell(row=row, column=3).value = gasto.categoria.nombre
        ws.cell(row=row, column=4).value = float(gasto.monto)
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
            if col == 4:
                ws.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # Ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    
    # Total
    total_row = len(gastos) + 2
    ws.cell(row=total_row, column=3).value = "TOTAL"
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4).value = f"=SUM(D2:D{len(gastos)+1})"
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=4).number_format = '$#,##0.00'
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'gastos_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/exportar/ingresos')
@login_required
def exportar_ingresos():

    ingresos = _filtrar_por_fecha(
        Ingreso.query.filter_by(empresa_id=current_user.empresa_id),
        Ingreso.fecha).order_by(Ingreso.fecha).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingresos"
    
    # Estilos
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    headers = ['Fecha', 'Descripción', 'Categoría', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, ingreso in enumerate(ingresos, 2):
        ws.cell(row=row, column=1).value = ingreso.fecha
        ws.cell(row=row, column=2).value = ingreso.descripcion
        ws.cell(row=row, column=3).value = ingreso.categoria.nombre
        ws.cell(row=row, column=4).value = float(ingreso.monto)
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
            if col == 4:
                ws.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # Ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    
    # Total
    total_row = len(ingresos) + 2
    ws.cell(row=total_row, column=3).value = "TOTAL"
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4).value = f"=SUM(D2:D{len(ingresos)+1})"
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=4).number_format = '$#,##0.00'
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'ingresos_{datetime.now().strftime("%Y%m%d")}.xlsx')

# ======================== FUNCIONES AUXILIARES ========================

def gastos_por_categoria(gastos):
    categorias_dict = {}
    for gasto in gastos:
        nombre = gasto.categoria.nombre
        if nombre not in categorias_dict:
            categorias_dict[nombre] = 0
        categorias_dict[nombre] += float(gasto.monto)
    return categorias_dict

def ingresos_por_categoria(ingresos):
    categorias_dict = {}
    for ingreso in ingresos:
        nombre = ingreso.categoria.nombre
        if nombre not in categorias_dict:
            categorias_dict[nombre] = 0
        categorias_dict[nombre] += float(ingreso.monto)
    return categorias_dict

@app.errorhandler(404)
def not_found(error):
    return render_template('404.html'), 404

@app.route('/cuenta/eliminar', methods=['GET', 'POST'])
@login_required
def eliminar_cuenta():
    if request.method == 'POST':
        try:
            # `usuario_id` se fue con la reasignacion de categorias: era su
            # unico uso (FASE-CATEGORIA-S1).
            empresa_id = current_user.empresa_id
            nombre_usuario = current_user.nombre
            
            # LA CAJA YA NO SE BORRA ACA (FASE-CAJA-GENERAL-S2).
            #
            # Estaban los dos `.delete()` de gasto e ingreso, mas el cascade
            # del modelo: borrarse la cuenta ponia en cero el libro de caja de
            # la empresa entera. Ahora las filas quedan, con `usuario_id` en
            # NULL (por eso la columna paso a nullable) y el `empresa_id`
            # intacto -- exactamente lo que FASE-AUDITORIA-S2 hizo con el
            # historial, y por el mismo motivo: en que se gasto la plata no es
            # un dato de la persona que lo tipeo.
            #
            # SQLAlchemy las anula solas, por los backrefs 'gastos'/'ingresos'.

            # LAS CATEGORIAS YA NO SE REASIGNAN ACA (FASE-CATEGORIA-S1).
            #
            # Estaba el parche que buscaba un "heredero" en la empresa y le
            # pasaba las categorias del que se iba (y las borraba si no quedaba
            # nadie). Hacia falta solo porque `categoria.usuario_id` era NOT
            # NULL con cascade: sin heredero, borrarse la cuenta se llevaba el
            # vocabulario y dejaba sin etiqueta a los gastos que la slice
            # anterior acababa de salvar. Ahora el duenio es `empresa_id`, que
            # nadie toca al borrar una persona: SQLAlchemy anula `usuario_id`
            # (por eso paso a nullable) y la categoria sigue viva.

            # EL HISTORIAL YA NO SE BORRA ACA (FASE-AUDITORIA-S2).
            #
            # Estaba `Historial.query.filter_by(usuario_id=...).delete()`: la
            # accion mas destructiva del sistema era tambien la que limpiaba su
            # propia evidencia. Ahora las filas quedan, con usuario_id en NULL
            # (por eso la columna paso a nullable) y el empresa_id intacto, asi
            # que los que siguen en la empresa ven lo que hizo el que se fue.
            #
            # SQLAlchemy las anula solo, por el backref 'historial'.

            # Eliminar usuario
            db.session.delete(current_user)

            # Eliminar empresa si no tiene otros usuarios
            empresa = Empresa.query.get(empresa_id)
            if empresa and Empresa.query.filter_by(id=empresa_id).first():
                otros_usuarios = Usuario.query.filter_by(empresa_id=empresa_id).count()
                if otros_usuarios == 0:
                    # Se va la empresa entera: no queda nadie a quien mostrarle
                    # este historial ni a quien servirle este vocabulario, y en
                    # las dos tablas `empresa_id` es NOT NULL, asi que dejarlas
                    # huerfanas volteria el borrado con un error de FK. Este es
                    # el unico caso en que el rastro se va.
                    #
                    # `categoria` entra en la lista desde FASE-CATEGORIA-S1: es
                    # la contracara de que ahora sobreviva a cualquier usuario.
                    Categoria.query.filter_by(empresa_id=empresa_id).delete()
                    Historial.query.filter_by(empresa_id=empresa_id).delete()
                    db.session.delete(empresa)

            db.session.commit()
            
            logout_user()
            flash('Tu cuenta ha sido eliminada correctamente', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            flash(f'Error al eliminar cuenta: {str(e)}', 'danger')
            return redirect(url_for('dashboard'))
    
    return render_template('eliminar_cuenta.html', usuario=current_user)

if __name__ == '__main__':
    print("\n" + "="*50)
    print("🔍 CONFIGURACIÓN DE EMAIL:")
    print(f"✓ MAIL_USERNAME: {app.config['MAIL_USERNAME']}")
    print(f"✓ MAIL_PASSWORD: {'*' * 20}")
    print(f"✓ MAIL_SERVER: {app.config['MAIL_SERVER']}")
    print("="*50 + "\n")
    print("Iniciando servidor...")
    print("Abre: http://localhost:5000")
    app.run(debug=True, host='127.0.0.1', port=5000)